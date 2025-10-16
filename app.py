import concurrent.futures
from flask import Flask, request, jsonify, send_file, render_template, redirect, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
import io
import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from database import db, User, UploadHistory, init_db, can_add_user

app = Flask(__name__)

# ========== CONFIGURAÇÕES ==========
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///hislogistica.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

init_db(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Faça login para acessar esta página.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Pega API Key do ambiente ou usa a padrão (MUDE NO RENDER!)
API_KEY = os.environ.get('GOOGLE_API_KEY', 'AIzaSyAbfZN5HQv_JX0nQTPoasOWt6iMWB3dvNE')

# ========== FUNÇÃO DE BUSCA ULTRA COMPLETA ==========
def buscar_empresa_google(nome_empresa):
    """
    Busca TODAS as informações disponíveis da empresa na Places API (New)
    """
    if not isinstance(nome_empresa, str):
        nome_empresa = str(nome_empresa)
    
    nome_empresa_corrigido = nome_empresa.strip().replace('  ', ' ')
    url = "https://places.googleapis.com/v1/places:searchText"
    
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": API_KEY,
        # FieldMask com TODOS os campos disponíveis
        "X-Goog-FieldMask": "places.id,places.displayName,places.formattedAddress,places.nationalPhoneNumber,places.internationalPhoneNumber,places.websiteUri,places.googleMapsUri,places.rating,places.userRatingCount,places.businessStatus,places.location,places.types,places.regularOpeningHours,places.priceLevel"
    }
    
    body = {
        "textQuery": nome_empresa_corrigido,
        "languageCode": "pt-BR"
    }
    
    try:
        response = requests.post(url, headers=headers, json=body, timeout=15)
        
        if response.status_code == 200:
            result = response.json()
            
            if 'places' in result and len(result['places']) > 0:
                place = result['places'][0]
                
                # Extrai TODAS as informações
                dados = {
                    'nome': place.get('displayName', {}).get('text', nome_empresa_corrigido),
                    'telefone_nacional': place.get('nationalPhoneNumber', 'Não encontrado'),
                    'telefone_internacional': place.get('internationalPhoneNumber', 'Não encontrado'),
                    'endereco': place.get('formattedAddress', 'Não encontrado'),
                    'website': place.get('websiteUri', 'Não encontrado'),
                    'google_maps_url': place.get('googleMapsUri', 'Não encontrado'),
                    'avaliacao': place.get('rating', 0),
                    'numero_avaliacoes': place.get('userRatingCount', 0),
                    'status': place.get('businessStatus', 'DESCONHECIDO'),
                    'latitude': place.get('location', {}).get('latitude', 0),
                    'longitude': place.get('location', {}).get('longitude', 0),
                    'tipos': ', '.join(place.get('types', [])),
                    'preco': '💰' * place.get('priceLevel', 0) if place.get('priceLevel') else 'N/A',
                    'horario': formatar_horario(place.get('regularOpeningHours', {}))
                }
                
                # Usa telefone nacional como principal
                telefone_principal = dados['telefone_nacional']
                if telefone_principal == 'Não encontrado':
                    telefone_principal = dados['telefone_internacional']
                
                print(f"✅ {dados['nome']} | {telefone_principal} | ⭐ {dados['avaliacao']}")
                
                return dados, telefone_principal
                
    except Exception as e:
        print(f"❌ Erro: {nome_empresa} - {e}")
    
    # Retorna dados vazios se não encontrou
    return {
        'nome': nome_empresa,
        'telefone_nacional': 'Não encontrado',
        'telefone_internacional': 'Não encontrado',
        'endereco': 'Não encontrado',
        'website': 'Não encontrado',
        'google_maps_url': 'Não encontrado',
        'avaliacao': 0,
        'numero_avaliacoes': 0,
        'status': 'NÃO ENCONTRADO',
        'latitude': 0,
        'longitude': 0,
        'tipos': 'N/A',
        'preco': 'N/A',
        'horario': 'Não disponível'
    }, 'Não encontrado'

def formatar_horario(horario_dict):
    """Formata o horário de funcionamento de forma legível"""
    if not horario_dict or 'weekdayDescriptions' not in horario_dict:
        return 'Horário não disponível'
    
    descricoes = horario_dict['weekdayDescriptions']
    return ' | '.join(descricoes[:3])  # Primeiros 3 dias

def buscar_empresas_em_paralelo(empresas):
    """Busca múltiplas empresas simultaneamente"""
    with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
        results = list(executor.map(buscar_empresa_google, empresas))
    return results

# ========== GERAÇÃO DE PLANILHA EXCEL COMPLETA ==========
def gerar_planilha_completa(file_path):
    """Gera planilha Excel com TODAS as informações"""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    empresas = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            empresas.append(row[0])

    print(f"🔄 Buscando {len(empresas)} empresas com informações COMPLETAS...")
    resultados = buscar_empresas_em_paralelo(empresas)

    # Filtra apenas empresas encontradas
    resultados_filtrados = [
        dados for dados, tel in resultados 
        if tel != 'Não encontrado'
    ]
    
    total = len(resultados)
    encontrados = len(resultados_filtrados)
    nao_encontrados = total - encontrados
    api_cost = total * 0.032
    
    print(f"\n📊 Total: {total} | ✅ Encontrados: {encontrados} | ❌ Removidos: {nao_encontrados}")
    print(f"💰 Custo: ${api_cost:.2f}\n")

    # Cria nova planilha
    novo_workbook = openpyxl.Workbook()
    novo_sheet = novo_workbook.active
    novo_sheet.title = 'Empresas Completas'

    # Estilos
    header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    content_fill_even = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    content_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Cabeçalhos COMPLETOS
    headers = [
        'Nome da Empresa', 'Telefone Nacional', 'Telefone Internacional',
        'Endereço Completo', 'Website', 'Google Maps', 'Avaliação ⭐',
        'Nº Avaliações', 'Status', 'Latitude', 'Longitude',
        'Tipos de Negócio', 'Faixa de Preço', 'Horário de Funcionamento'
    ]
    
    novo_sheet.append(headers)
    
    # Formata cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = novo_sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Adiciona dados
    for row_num, dados in enumerate(resultados_filtrados, start=2):
        row_data = [
            dados['nome'],
            dados['telefone_nacional'],
            dados['telefone_internacional'],
            dados['endereco'],
            dados['website'],
            dados['google_maps_url'],
            dados['avaliacao'],
            dados['numero_avaliacoes'],
            dados['status'],
            dados['latitude'],
            dados['longitude'],
            dados['tipos'],
            dados['preco'],
            dados['horario']
        ]
        
        novo_sheet.append(row_data)
        
        # Aplica estilos
        for col_num in range(1, len(headers) + 1):
            cell = novo_sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            cell.fill = content_fill_even if row_num % 2 == 0 else content_fill_odd

    # Ajusta largura das colunas
    column_widths = [30, 20, 20, 50, 35, 35, 12, 15, 15, 15, 15, 30, 15, 60]
    for idx, width in enumerate(column_widths, 1):
        novo_sheet.column_dimensions[get_column_letter(idx)].width = width
    
    # Congela primeira linha
    novo_sheet.freeze_panes = 'A2'

    stream = io.BytesIO()
    novo_workbook.save(stream)
    stream.seek(0)
    
    return stream, resultados_filtrados, {
        'total': total,
        'encontrados': encontrados,
        'nao_encontrados': nao_encontrados,
        'api_cost': api_cost
    }

# ========== GERAÇÃO DE PDF ABSURDO ==========
def gerar_pdf_foda(dados_empresas, filename='relatorio_empresas.pdf'):
    """Gera PDF PROFISSIONAL com todas as informações"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo customizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1  # Centralizado
    )
    
    # Título
    elements.append(Paragraph("📊 RELATÓRIO COMPLETO DE EMPRESAS", title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Estatísticas
    stats_data = [
        ['📈 ESTATÍSTICAS DO RELATÓRIO'],
        [f'Total de Empresas: {len(dados_empresas)}'],
        [f'Data/Hora: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}']
    ]
    
    stats_table = Table(stats_data, colWidths=[10*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabela de empresas (dividida em páginas se necessário)
    chunk_size = 8  # Empresas por página
    
    for i in range(0, len(dados_empresas), chunk_size):
        chunk = dados_empresas[i:i+chunk_size]
        
        # Cabeçalho da tabela
        table_data = [[
            'Empresa', 'Telefone', 'Endereço', 'Website', 
            '⭐ Nota', 'Status', 'Tipos'
        ]]
        
        # Adiciona dados
        for empresa in chunk:
            table_data.append([
                Paragraph(empresa['nome'][:40], styles['Normal']),
                empresa['telefone_nacional'],
                Paragraph(empresa['endereco'][:50], styles['Normal']),
                Paragraph(empresa['website'][:30], styles['Normal']) if empresa['website'] != 'Não encontrado' else 'N/A',
                f"{empresa['avaliacao']:.1f}⭐",
                'ABERTO' if empresa['status'] == 'OPERATIONAL' else 'FECHADO',
                Paragraph(empresa['tipos'][:30], styles['Normal'])
            ])
        
        table = Table(table_data, colWidths=[1.8*inch, 1.2*inch, 2*inch, 1.5*inch, 0.8*inch, 0.9*inch, 1.8*inch])
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Conteúdo
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Bordas
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        
        # Page break se não for o último chunk
        if i + chunk_size < len(dados_empresas):
            elements.append(PageBreak())
    
    # Rodapé
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph("© 2025 HIS Logística - Todos os direitos reservados", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# ========== ROTAS DE AUTENTICAÇÃO (SEM MUDANÇAS) ==========
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            if not user.is_active:
                flash('Sua conta está desativada.', 'error')
                return redirect(url_for('login'))
            
            login_user(user)
            user.last_login = datetime.utcnow()
            db.session.commit()
            flash(f'Bem-vindo, {user.username}!', 'success')
            
            if user.is_admin:
                return redirect(url_for('admin_panel'))
            return redirect(url_for('index'))
        else:
            flash('Usuário ou senha incorretos.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=current_user)

# ========== UPLOAD E DOWNLOAD ==========
@app.route('/upload_planilha', methods=['POST'])
@login_required
def upload_planilha():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']
    
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    
    file_path = os.path.join('uploads', file.filename)
    file.save(file_path)

    try:
        planilha, dados_empresas, stats = gerar_planilha_completa(file_path)
        os.remove(file_path)
        
        # Salva planilha Excel
        os.makedirs('temp', exist_ok=True)
        excel_path = 'temp/planilha_atualizada.xlsx'
        with open(excel_path, 'wb') as f:
            f.write(planilha.getbuffer())
        
        # Gera PDF
        pdf_buffer = gerar_pdf_foda(dados_empresas)
        pdf_path = 'temp/relatorio_empresas.pdf'
        with open(pdf_path, 'wb') as f:
            f.write(pdf_buffer.getvalue())
        
        # Salva no histórico
        upload = UploadHistory(
            user_id=current_user.id,
            filename=file.filename,
            total_empresas=stats['total'],
            empresas_encontradas=stats['encontrados'],
            empresas_nao_encontradas=stats['nao_encontrados'],
            api_cost=stats['api_cost']
        )
        db.session.add(upload)
        db.session.commit()
        
        return jsonify({
            "link_excel": "/download_excel",
            "link_pdf": "/download_pdf",
            "stats": stats
        })
    except Exception as e:
        print(f"❌ Erro: {e}")
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': f'Erro: {str(e)}'}), 500

@app.route('/download_excel')
@login_required
def download_excel():
    return send_file('temp/planilha_atualizada.xlsx', 
                     as_attachment=True, 
                     download_name=f'empresas_completas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@app.route('/download_pdf')
@login_required
def download_pdf():
    return send_file('temp/relatorio_empresas.pdf', 
                     as_attachment=True,
                     download_name=f'relatorio_empresas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf')

# ========== ADMIN (SEM MUDANÇAS) ==========
@app.route('/admin')
@login_required
def admin_panel():
    if not current_user.is_admin:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))
    
    users = User.query.all()
    uploads = UploadHistory.query.order_by(UploadHistory.upload_date.desc()).limit(50).all()
    return render_template('admin.html', users=users, uploads=uploads)

@app.route('/admin/add_user', methods=['POST'])
@login_required
def add_user():
    if not current_user.is_admin:
        return jsonify({'error': 'Acesso negado'}), 403
    
    if not can_add_user():
        return jsonify({'error': 'Limite de 10 usuários atingido'}), 400
    
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Usuário já existe'}), 400
    
    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email já cadastrado'}), 400
    
    user = User(username=username, email=email, is_admin=False)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Usuário {username} criado!'})

@app.route('/admin/toggle_user/<int:user_id>', methods=['POST'])
@login_required
def toggle_user(user_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Acesso negado'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.is_admin:
        return jsonify({'error': 'Não pode desativar admin'}), 400
    
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'ativado' if user.is_active else 'desativado'
    return jsonify({'success': True, 'message': f'Usuário {status}!', 'is_active': user.is_active})

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        return jsonify({'error': 'Acesso negado'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.is_admin:
        return jsonify({'error': 'Não pode deletar admin'}), 400
    
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Usuário deletado!'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)